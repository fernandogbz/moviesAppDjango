from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Avg
from django.http import HttpResponse
from .models import Pelicula, Director, Genero, Reseña
from .forms import PeliculaForm, DirectorForm, GeneroForm, ReseñaForm, CustomUserCreationForm, PeliculaFilterForm
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create your views here.

def is_admin(user):
    return user.is_staff or user.is_superuser

# Vistas de autenticación
def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Cuenta creada exitosamente!')
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

# Vista principal
def home(request):
    # Obtener películas mejor calificadas
    peliculas_destacadas = Pelicula.objects.annotate(
        avg_rating=Avg('reseñas__calificacion')
    ).order_by('-avg_rating')[:6]
    
    context = {
        'peliculas_destacadas': peliculas_destacadas
    }
    return render(request, 'moviesApp/home.html', context)

# Vistas de películas
def pelicula_list(request):
    form = PeliculaFilterForm(request.GET)
    peliculas = Pelicula.objects.all()
    
    if form.is_valid():
        titulo = form.cleaned_data.get('titulo')
        genero = form.cleaned_data.get('genero')
        director = form.cleaned_data.get('director')
        año_desde = form.cleaned_data.get('año_desde')
        año_hasta = form.cleaned_data.get('año_hasta')
        
        if titulo:
            peliculas = peliculas.filter(titulo__icontains=titulo)
        if genero:
            peliculas = peliculas.filter(generos=genero)
        if director:
            peliculas = peliculas.filter(directores=director)
        if año_desde:
            peliculas = peliculas.filter(fechaEstreno__year__gte=año_desde)
        if año_hasta:
            peliculas = peliculas.filter(fechaEstreno__year__lte=año_hasta)
    
    # Paginación
    paginator = Paginator(peliculas, 6)  # 6 películas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form
    }
    return render(request, 'moviesApp/pelicula_list.html', context)

def pelicula_detail(request, pk):
    pelicula = get_object_or_404(Pelicula, pk=pk)
    reseñas = pelicula.reseñas.all().order_by('-fecha')
    user_reseña = None
    
    if request.user.is_authenticated:
        user_reseña = reseñas.filter(usuario=request.user).first()
    
    context = {
        'pelicula': pelicula,
        'reseñas': reseñas,
        'user_reseña': user_reseña
    }
    return render(request, 'moviesApp/pelicula_detail.html', context)

@user_passes_test(is_admin)
def pelicula_create(request):
    if request.method == 'POST':
        form = PeliculaForm(request.POST, request.FILES)
        if form.is_valid():
            pelicula = form.save()
            messages.success(request, 'Película creada exitosamente!')
            return redirect('pelicula_detail', pk=pelicula.pk)
    else:
        form = PeliculaForm()
    
    context = {
        'form': form,
        'title': 'Crear Película'
    }
    return render(request, 'moviesApp/pelicula_form.html', context)

@user_passes_test(is_admin)
def pelicula_edit(request, pk):
    pelicula = get_object_or_404(Pelicula, pk=pk)
    
    if request.method == 'POST':
        form = PeliculaForm(request.POST, request.FILES, instance=pelicula)
        if form.is_valid():
            pelicula = form.save()
            messages.success(request, 'Película actualizada exitosamente!')
            return redirect('pelicula_detail', pk=pelicula.pk)
    else:
        form = PeliculaForm(instance=pelicula)
    
    context = {
        'form': form,
        'title': 'Editar Película',
        'pelicula': pelicula
    }
    return render(request, 'moviesApp/pelicula_form.html', context)

@user_passes_test(is_admin)
def pelicula_delete(request, pk):
    pelicula = get_object_or_404(Pelicula, pk=pk)
    if request.method == 'POST':
        pelicula.delete()
        messages.success(request, 'Película eliminada exitosamente!')
        return redirect('pelicula_list')
    return render(request, 'moviesApp/pelicula_confirm_delete.html', {'pelicula': pelicula})

# Vistas de directores
@user_passes_test(is_admin)
def director_list(request):
    directores = Director.objects.all()
    # paginator = Paginator(directores, 10)
    # page_number = request.GET.get('page')
    # page_obj = paginator.get_page(page_number)
    # return render(request, 'moviesApp/director_list.html', {'page_obj': page_obj})
    context = {
        'directores': directores
    }
    return render(request, 'moviesApp/director_list.html', context)

@user_passes_test(is_admin)
def director_create(request):
    if request.method == 'POST':
        form = DirectorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Director creado exitosamente!')
            return redirect('director_list')
    else:
        form = DirectorForm()
    return render(request, 'moviesApp/director_form.html', {'form': form, 'title': 'Crear Director'})

@user_passes_test(is_admin)
def director_edit(request, pk):
    director = get_object_or_404(Director, pk=pk)
    if request.method == 'POST':
        form = DirectorForm(request.POST, instance=director)
        if form.is_valid():
            form.save()
            messages.success(request, 'Director actualizado exitosamente!')
            return redirect('director_list')
    else:
        form = DirectorForm(instance=director)
    return render(request, 'moviesApp/director_form.html', {'form': form, 'title': 'Editar Director', 'director': director})

@user_passes_test(is_admin)
def director_delete(request, pk):
    director = get_object_or_404(Director, pk=pk)
    
    if request.method == 'POST':
        director.delete()
        messages.success(request, 'Director eliminado exitosamente!')
        return redirect('director_list')
    
    context = {
        'director': director
    }
    return render(request, 'moviesApp/director_confirm_delete.html', context)

# Vistas de géneros
@user_passes_test(is_admin)
def genero_list(request):
    generos = Genero.objects.all()
    context = {
        'generos': generos  
    }
    return render(request, 'moviesApp/genero_list.html', context)

@user_passes_test(is_admin)
def genero_create(request):
    if request.method == 'POST':
        form = GeneroForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Género creado exitosamente!')
            return redirect('genero_list')
    else:
        form = GeneroForm()
    return render(request, 'moviesApp/genero_form.html', {'form': form, 'title': 'Crear Género'})

@user_passes_test(is_admin)
def genero_edit(request, pk):
    genero = get_object_or_404(Genero, pk=pk)
    if request.method == 'POST':
        form = GeneroForm(request.POST, instance=genero)
        if form.is_valid():
            form.save()
            messages.success(request, 'Género actualizado exitosamente!')
            return redirect('genero_list')
    else:
        form = GeneroForm(instance=genero)
    return render(request, 'moviesApp/genero_form.html', {'form': form, 'title': 'Editar Género', 'genero': genero})

@user_passes_test(is_admin)
def genero_delete(request, pk):
    genero = get_object_or_404(Genero, pk=pk)
    if request.method == 'POST':
        genero.delete()
        messages.success(request, 'Género eliminado exitosamente!')
        return redirect('genero_list')
    return render(request, 'moviesApp/genero_confirm_delete.html', {'genero': genero})

# Vistas de reseñas
@login_required
def reseña_create(request, pelicula_pk):
    pelicula = get_object_or_404(Pelicula, pk=pelicula_pk)
    
    # Verificar si el usuario ya ha dejado una reseña
    existing_review = Reseña.objects.filter(pelicula=pelicula, usuario=request.user).first()
    if existing_review:
        messages.warning(request, 'Ya has dejado una reseña para esta película.')
        return redirect('pelicula_detail', pk=pelicula_pk)
    
    if request.method == 'POST':
        form = ReseñaForm(request.POST)
        if form.is_valid():
            reseña = form.save(commit=False)
            reseña.pelicula = pelicula
            reseña.usuario = request.user
            reseña.save()
            messages.success(request, 'Reseña creada exitosamente!')
            return redirect('pelicula_detail', pk=pelicula_pk)
    else:
        form = ReseñaForm()
    
    return render(request, 'moviesApp/reseña_form.html', {
        'form': form, 
        'pelicula': pelicula, 
        'title': 'Crear Reseña'
    })

@user_passes_test(is_admin)
def reseña_list(request):
    reseñas = Reseña.objects.all().order_by('-fecha')
    paginator = Paginator(reseñas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'moviesApp/reseña_list.html', {'page_obj': page_obj})

# Vista de ranking
def ranking(request):
    peliculas = Pelicula.objects.annotate(
        avg_rating=Avg('reseñas__calificacion')
    ).filter(avg_rating__isnull=False).order_by('-avg_rating')
    
    paginator = Paginator(peliculas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'moviesApp/ranking.html', {'page_obj': page_obj})

# Vista de exportación a Excel
def export_peliculas_excel(request):
    # Crear workbook y worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Películas'
    
    # Headers
    headers = ['Título', 'Descripción', 'Fecha Estreno', 'Duración (min)', 
               'Directores', 'Géneros', 'Calificación Promedio', 'Total Reseñas']
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = header
    
    # Data
    peliculas = Pelicula.objects.all().annotate(avg_rating=Avg('reseñas__calificacion'))
    
    for row_num, pelicula in enumerate(peliculas, 2):
        worksheet[f'A{row_num}'] = pelicula.titulo
        worksheet[f'B{row_num}'] = pelicula.descripcion
        worksheet[f'C{row_num}'] = pelicula.fechaEstreno.strftime('%Y-%m-%d')
        worksheet[f'D{row_num}'] = pelicula.duracion
        worksheet[f'E{row_num}'] = ', '.join([d.nombre for d in pelicula.directores.all()])
        worksheet[f'F{row_num}'] = ', '.join([g.nombre for g in pelicula.generos.all()])
        worksheet[f'G{row_num}'] = round(pelicula.avg_rating or 0, 2)
        worksheet[f'H{row_num}'] = pelicula.get_total_reviews()
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="peliculas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response
